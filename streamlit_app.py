import streamlit as st
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# -----------------------
# App Title and Intro
# -----------------------
st.set_page_config(page_title="Leadership on the Line", page_icon="✈️", layout="centered")
st.title("Leadership on the Line ✈️")
st.markdown("Welcome to the inspection tracker — enter your details below to begin.")
st.write("❌ Fields are required to save document")

# -----------------------
# User Inputs
# -----------------------
who = st.text_input("❌ Who is inspecting?  (Enter Last name)")
aircraft_number = st.text_input("❌ Location of Report")
inspection_time = st.text_input("❌ Time of Inspection (Military time)")
st.write("Not All boxes are required for submission")
line_badge = st.selectbox("Do individuals have their line badge?", ["Yes", "No"])
badge_showing = st.selectbox("Is it showing?", ["Yes", "No"])
ppe = st.selectbox("PPE Worn Correctly?", ["Yes", "No", "N/A"])

st.markdown("---")
st.subheader("Score the following (1–5, with 5 being best):")

cleanliness = st.text_input("Cleanliness Inside/Outside")
safe_maint = st.text_input("Safe For Maintenance")
organized_storage = st.text_input("Organized Cargo/Storage")
organized_flightdeck = st.text_input("Organized Flight Deck")
forms = st.text_input("Forms 781s current/accurate/signed?")
fod_check = st.text_input("FOD Check?")
age_pos = st.text_input("AGE Positioned Safely?")
comments = st.text_area("Additional Comments", placeholder="Type 'No' if none")

# -----------------------
# Save Button
# -----------------------
if st.button("💾 Save Inspection"):
    if not who or not aircraft_number or not inspection_time:
        st.error("Please fill out required fields: Name, Aircraft Number, and Inspection Time.")
    else:
        # Format data
        date_now = datetime.now().strftime("%Y-%m-%d %H:%M")
        who_upper = who.upper()
        line_badge = line_badge.upper()
        badge_showing = badge_showing.upper()
        ppe = ppe.upper()

        filename = "Leadership_on_line.xlsx"
        save_path = os.path.join(os.getcwd(), filename)

        # Create the question/answer structure
        inspection_data = [
            ("Date of Inspection", date_now),
            ("Who spectated?", who_upper),
            ("What Aircraft?", aircraft_number),
            ("Time inspection was done", inspection_time),
            ("Line Badge?", line_badge),
            ("Showing?", badge_showing),
            ("PPE worn correctly?", ppe),
            ("Cleanliness Inside/Outside?", cleanliness),
            ("Safe For Maintenance?", safe_maint),
            ("Organized Cargo/Storage?", organized_storage),
            ("Organized Flight Deck?", organized_flightdeck),
            ("Forms?", forms),
            ("FOD Check?", fod_check),
            ("AGE Positioned correctly?", age_pos),
            ("Comments", comments),
            ("--- NEW ---", ""),
        ]

        from openpyxl import Workbook, load_workbook

        # -----------------------
        # Robust Excel Save Logic
        # -----------------------
        if os.path.exists(save_path):
            try:
                book = load_workbook(save_path)
                sheet = book.active
            except Exception as e:
                # File is corrupted or not a valid Excel; rebuild it
                st.warning("⚠️ Existing Excel file was corrupted. Recreating it now.")
                os.remove(save_path)
                book = Workbook()
                sheet = book.active
                sheet.title = "Inspections"
                sheet.append(["Question", "Answer"])

            # Append data rows
            for q, a in inspection_data:
                sheet.append([q, a])
            book.save(save_path)

        else:
            # Create new workbook from scratch
            book = Workbook()
            sheet = book.active
            sheet.title = "Inspections"
            sheet.append(["Question", "Answer"])
            for q, a in inspection_data:
                sheet.append([q, a])
            book.save(save_path)

        st.success("✅ Inspection saved successfully!")
        st.info(f"File saved at: {save_path}")


# -----------------------
# Display File (Excel)
# -----------------------
if st.button("📄 View Saved Inspections"):
    save_path = os.path.join(os.getcwd(), "Leadership_on_line.xlsx")

    if os.path.exists(save_path):
        try:
            data = pd.read_excel(save_path, sheet_name="Inspections", names=["Question", "Answer"])
            st.subheader("📋 Saved Inspections Log")
            st.dataframe(data, use_container_width=True)

            # Download Excel file button
            with open(save_path, "rb") as f:
                st.download_button(
                    label="⬇️ Download Excel File",
                    data=f,
                    file_name="Leadership_on_line.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        except Exception as e:
            st.error(f"⚠️ Error reading saved inspections: {e}")
    else:
        st.warning("No saved inspections found yet.")
